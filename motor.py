# -*- coding: utf-8 -*-
"""
Motor Control ORAMI
- Lee el buzón transacciones@nacionalissimo.com (IMAP)
- Correo del banco (Banorte, transferencia SPEI) -> crea recarga PENDIENTE + manda notificacion push
- Correo de ORAMI (xlsx estado de cuenta) -> agrega/actualiza movimientos y marca recargas VERIFICADAS por clave de rastreo
Se ejecuta en GitHub Actions (cron).
"""
import os, ssl, imaplib, email, re, json, io, zipfile, smtplib, socket
socket.setdefaulttimeout(90)   # ninguna operacion de red se cuelga para siempre, html
import xml.etree.ElementTree as ET
from email.header import decode_header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime, timedelta

import firebase_admin
from firebase_admin import credentials, firestore, messaging

# ---------- Config desde variables de entorno (GitHub Secrets) ----------
IMAP_HOST = os.environ.get("IMAP_HOST", "mail.nacionalissimo.com")
IMAP_PORT = int(os.environ.get("IMAP_PORT", "993"))
IMAP_USER = os.environ["IMAP_USER"]
IMAP_PASS = os.environ["IMAP_PASS"]
SA_JSON   = os.environ["FIREBASE_SA"]   # contenido JSON de la cuenta de servicio
SMTP_HOST = os.environ.get("SMTP_HOST", "mail.nacionalissimo.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "465"))
REENVIO_BBVA = os.environ.get("REENVIO_BBVA", "jgortizm@hotmail.com")  # a quien se reenvia el aviso de BBVA

SALDO_INICIAL = 23833.71  # saldo real al 1-jun-2026 (reconciliado con la lista completa de recargas + estado de cuenta ORAMI 30-jun-2026; no incluye el retenido). Se ajustara con el ciclo mensual.

# ---------- Firebase ----------
cred = credentials.Certificate(json.loads(SA_JSON))
firebase_admin.initialize_app(cred)
db = firestore.client()

def log(*a): print("[motor]", *a, flush=True)

def comision(t): return round((t/1.16)*0.055, 2)
def abono_neto(t): return round(t - comision(t), 2)

# ---------- CIERRE MENSUAL ----------
# Al iniciar cada mes: genera el estado de cuenta del mes anterior (Excel), lo manda al
# correo, borra los movimientos VERIFICADOS archivados y arrastra el saldo final como
# saldo inicial del mes nuevo (config/cierre, que la app lee en vivo). Los PENDIENTES se
# pasan al mes nuevo sin archivar. Los ids archivados se guardan en config/archivados
# para que un reporte reenviado de ORAMI no los resucite.
BOOTSTRAP_SALDO = 23833.71   # saldo real al 1-jun-2026 (reconciliado con ORAMI)
BOOTSTRAP_MES   = "2026-06"  # primer mes administrado por el cierre

MESES_NOMBRE = ["","Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio",
                "Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
ARCHIVADOS = set()   # ids ya archivados en cierres previos (no re-crearlos jamas)

def ahora_mx():
    """Hora de Mexico (UTC-6 fijo); los runners de GitHub estan en UTC."""
    return datetime.utcnow() - timedelta(hours=6)

def _serial(dt): return (dt - datetime(1899,12,30)).total_seconds()/86400.0
def _mes_inicio(mes): a,m = mes.split("-"); return datetime(int(a), int(m), 1)
def _mes_siguiente(mes):
    a, m = int(mes[:4]), int(mes[5:7]) + 1
    if m == 13: a += 1; m = 1
    return "%04d-%02d" % (a, m)
def _mes_titulo(mes): return "%s %s" % (MESES_NOMBRE[int(mes[5:7])], mes[:4])

def cargar_cierre():
    ref = db.collection("config").document("cierre")
    s = ref.get()
    if s.exists: return s.to_dict()
    datos = {"saldoInicial": BOOTSTRAP_SALDO, "mesActual": BOOTSTRAP_MES,
             "fechaCierre": _serial(_mes_inicio(BOOTSTRAP_MES)),
             "actualizado": ahora_mx().strftime("%Y-%m-%d %H:%M:%S")}
    ref.set(datos)
    log("config/cierre inicializado:", datos)
    return datos

def cargar_archivados():
    global ARCHIVADOS
    try:
        s = db.collection("config").document("archivados").get()
        ARCHIVADOS = set((s.to_dict() or {}).get("ids", [])) if s.exists else set()
    except Exception as e:
        log("no se pudieron cargar archivados:", e); ARCHIVADOS = set()

def ya_archivado(doc_id):
    if doc_id in ARCHIVADOS:
        log("omitido (pertenece a un mes ya cerrado):", doc_id); return True
    return False

def _excel_mes(archivar, mes, saldo_ini, tot_ab, tot_ca, saldo_fin):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook(); ws = wb.active; ws.title = mes
    neg = Font(bold=True)
    ws.append(["Control ORAMI - Estado de cuenta", _mes_titulo(mes)]); ws["A1"].font = neg
    ws.append(["Saldo inicial del mes", saldo_ini])
    ws.append([])
    ws.append(["Fecha","Hora","Tipo","Servicio","Detalle","Recibo","Recibo ORAMI",
               "Transferencia","Comisión","Abono neto","Cargo","Estado"])
    for c in ws[4]: c.font = neg
    for d in sorted(archivar, key=lambda x: float(x.to_dict().get("orden",0))):
        m = d.to_dict()
        if m.get("tipo") == "abono":
            t = float(m.get("transferencia",0) or 0)
            ws.append([m.get("fechaFull",""), m.get("hora",""), "Recarga", m.get("servicio",""),
                       m.get("banco",""), str(m.get("recibo","")), str(m.get("reciboOrami","")),
                       t, comision(t), abono_neto(t), "", m.get("estado","")])
        else:
            ws.append([m.get("fechaFull",""), m.get("hora",""), "Cargo", m.get("servicio",""),
                       m.get("banco",""), str(m.get("recibo","")), str(m.get("reciboOrami","")),
                       "", "", "", float(m.get("monto",0) or 0), m.get("estado","")])
    ws.append([])
    ws.append(["","","","","","","Total recargas netas", "", "", tot_ab, "", ""])
    ws.append(["","","","","","","Total cargos", "", "", "", tot_ca, ""])
    ws.append(["","","","","","","SALDO FINAL DEL MES", "", "", saldo_fin, "", ""])
    for fila in (ws.max_row-2, ws.max_row-1, ws.max_row):
        for c in ws[fila]: c.font = neg
    anchos = [12,10,9,20,38,16,16,14,11,12,12,11]
    for i,a in enumerate(anchos,1):
        ws.column_dimensions[ws.cell(row=4,column=i).column_letter].width = a
    for fila in ws.iter_rows(min_row=2):
        for c in fila:
            if isinstance(c.value,(int,float)): c.number_format = "#,##0.00"
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()

def enviar_cierre(mes, adjunto, nombre, resumen):
    """Manda el estado de cuenta del mes al buzon. Si falla LANZA excepcion -> el cierre
    se aborta y se reintenta en la siguiente corrida (nunca se borra sin respaldo enviado)."""
    msg = MIMEMultipart()
    msg["Subject"] = "Cierre ORAMI - %s" % _mes_titulo(mes)
    msg["From"] = IMAP_USER; msg["To"] = IMAP_USER
    msg.attach(MIMEText(resumen, "plain", "utf-8"))
    adj = MIMEApplication(adjunto, Name=nombre)
    adj["Content-Disposition"] = 'attachment; filename="%s"' % nombre
    msg.attach(adj)
    s = _conectar_smtp()
    s.sendmail(IMAP_USER, [IMAP_USER], msg.as_string())
    s.quit()
    log("estado de cuenta de %s enviado al buzon" % mes)

def cierre_mensual():
    global ARCHIVADOS
    cfg = cargar_cierre()
    hoy_mes = ahora_mx().strftime("%Y-%m")
    while cfg.get("mesActual","") < hoy_mes:
        mes = cfg["mesActual"]; sig = _mes_siguiente(mes)
        corte = _serial(_mes_inicio(sig))
        todos = list(db.collection("movimientos").stream())
        archivar = []
        for d in todos:
            m = d.to_dict()
            if m.get("tipo") not in ("abono","cargo"): continue
            if m.get("estado") != "verificada": continue          # pendientes pasan al mes nuevo
            if float(m.get("orden",0) or 0) >= corte: continue    # lo del mes nuevo se queda
            archivar.append(d)
        tot_ab = round(sum(abono_neto(float(d.to_dict().get("transferencia",0) or 0))
                           for d in archivar if d.to_dict().get("tipo")=="abono"), 2)
        tot_ca = round(sum(float(d.to_dict().get("monto",0) or 0)
                           for d in archivar if d.to_dict().get("tipo")=="cargo"), 2)
        saldo_ini = round(float(cfg.get("saldoInicial",0)), 2)
        saldo_fin = round(saldo_ini + tot_ab - tot_ca, 2)
        resumen = ("Cierre del mes %s\n\n"
                   "   Saldo inicial:        $%s\n"
                   "   + Recargas netas:     $%s\n"
                   "   - Cargos:             $%s\n"
                   "   = SALDO FINAL:        $%s   (saldo inicial de %s)\n\n"
                   "Movimientos archivados: %d (ver Excel adjunto).\n"
                   "Los movimientos pendientes de confirmar por ORAMI pasan al mes nuevo.\n\n"
                   "Mensaje automático del motor Control ORAMI.") % (
                   _mes_titulo(mes), "{:,.2f}".format(saldo_ini), "{:,.2f}".format(tot_ab),
                   "{:,.2f}".format(tot_ca), "{:,.2f}".format(saldo_fin), _mes_titulo(sig), len(archivar))
        excel = _excel_mes(archivar, mes, saldo_ini, tot_ab, tot_ca, saldo_fin)
        enviar_cierre(mes, excel, "Cierre_ORAMI_%s.xlsx" % mes, resumen)   # si falla, aborta aqui
        # correo enviado -> ahora si: registrar ids archivados y borrar
        nuevos_ids = {d.id for d in archivar}
        cargar_archivados()
        db.collection("config").document("archivados").set({"ids": sorted(ARCHIVADOS | nuevos_ids)})
        ARCHIVADOS |= nuevos_ids
        for d in archivar: d.reference.delete()
        cfg = {"saldoInicial": saldo_fin, "mesActual": sig, "fechaCierre": corte,
               "actualizado": ahora_mx().strftime("%Y-%m-%d %H:%M:%S"),
               "ultimoCierre": {"mes": mes, "saldoInicial": saldo_ini, "abonosNetos": tot_ab,
                                 "cargos": tot_ca, "saldoFinal": saldo_fin, "movimientos": len(archivar)}}
        db.collection("config").document("cierre").set(cfg)
        log("CIERRE %s: %d movimientos archivados, saldo final %.2f" % (mes, len(archivar), saldo_fin))
        enviar_push("Cierre de mes",
                    "Se cerró %s con saldo de $%s. El estado de cuenta quedó en tu correo. %s inicia con ese saldo." %
                    (_mes_titulo(mes), "{:,.2f}".format(saldo_fin), _mes_titulo(sig)))
    return cfg

# ---------- Utilidades de correo ----------
def decode(s):
    if not s: return ""
    parts = decode_header(s); out = ""
    for txt, enc in parts:
        out += txt.decode(enc or "utf-8", "ignore") if isinstance(txt, bytes) else txt
    return out

def get_body(msg):
    """Devuelve el texto plano del correo (quitando HTML si hace falta)."""
    body = ""
    if msg.is_multipart():
        for part in msg.walk():
            ct = part.get_content_type()
            if ct in ("text/plain", "text/html") and "attachment" not in str(part.get("Content-Disposition")):
                try: body += part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8", "ignore")
                except Exception: pass
    else:
        try: body = msg.get_payload(decode=True).decode(msg.get_content_charset() or "utf-8", "ignore")
        except Exception: body = str(msg.get_payload())
    body = re.sub(r"(?is)<(style|script)\b[^>]*>.*?</\1>", " ", body)  # quitar CSS/JS embebido
    body = re.sub(r"<[^>]+>", " ", body)   # quitar tags HTML
    body = html.unescape(body)   # decodificar entidades: &#36;->$, &oacute;->o, &nbsp;->espacio
    return re.sub(r"\s+", " ", body)

def get_xlsx_attachment(msg):
    for part in msg.walk():
        fn = decode(part.get_filename())
        if fn and fn.lower().endswith(".xlsx"):
            return part.get_payload(decode=True)
    return None

# ---------- Notificaciones push ----------
def enviar_push(titulo, texto):
    tokens = [d.to_dict().get("token") for d in db.collection("tokens").stream()]
    tokens = [t for t in tokens if t]
    enviados = 0
    for t in tokens:
        try:
            messaging.send(messaging.Message(
                notification=messaging.Notification(title=titulo, body=texto),
                token=t,
                webpush=messaging.WebpushConfig(notification=messaging.WebpushNotification(icon="icon.svg"))
            ))
            enviados += 1
        except Exception as e:
            log("push fallo:", e)
    log(f"push enviado a {enviados}/{len(tokens)} dispositivos")

# ---------- Procesar correo del BANCO (transferencia) ----------
def procesar_banco(body):
    importe = re.search(r"Importe a Transferir:\s*\$?\s*([\d,]+\.\d{2})", body)
    clave   = re.search(r"Clave de Rastreo:\s*([A-Z0-9]+)", body)
    fechah  = re.search(r"Fecha y Hora de Operaci[oó]n:\s*([0-9]{1,2}/\w+/[0-9]{4}[^A-Za-z]*\d{2}:\d{2}:\d{2})", body)
    ref     = re.search(r"N[uú]mero de Referencia:\s*(\d+)", body)
    benef   = re.search(r"Nombre del Beneficiario:\s*([^N]+?)\s+(?:CLABE|RFC)", body)
    if not (importe and clave):
        log("correo banco sin importe/clave, ignorado"); return
    monto = float(importe.group(1).replace(",", ""))
    cve = clave.group(1)
    doc_id = "rec-" + cve
    if ya_archivado(doc_id): return
    ref_doc = db.collection("movimientos").document(doc_id)
    if ref_doc.get().exists:
        log("transferencia ya registrada:", cve); return
    fh = fechah.group(1).strip() if fechah else ""
    try:
        dt = datetime.strptime(re.sub(r"\s+", " ", fh)[:20], "%d/%b/%Y %H:%M:%S")
    except Exception:
        dt = datetime.now()
    orden = (dt - datetime(1899,12,30)).total_seconds()/86400.0
    MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    ref_doc.set({
        "tipo":"abono","servicio":"Recarga (transferencia)",
        "fecha":"%d-%s"%(dt.day, MESES[dt.month-1]),
        "fechaFull":"%d-%s-%d"%(dt.day, MESES[dt.month-1], dt.year),
        "hora":dt.strftime("%H:%M:%S"),"recibo":ref.group(1) if ref else "",
        "banco":"SPEI a ORAMI - Clave "+cve,
        "transferencia":monto,"estado":"pendiente","orden":orden,
        "comprobante":{
            "banco":"Banorte en su Empresa","operacion":"Transferencia SPEI",
            "fechaHora":fh,"ordenante":"MIGUEL ANGEL VALLEJO FLORES",
            "beneficiario":(benef.group(1).strip() if benef else "CONSULTORIA EMPRESARIAL ORAMI S.C."),
            "clabe":"","bancoDestino":"BAJIO","importe":"$%s MN"%importe.group(1),
            "referencia":ref.group(1) if ref else "","concepto":"ORAMI",
            "aplicacion":fh.split(" ")[0] if fh else "","clave":cve
        }
    })
    log("recarga creada:", cve, monto)
    enviar_push("Transferencia registrada", "Transferiste $%s a ORAMI. Recarga pendiente de confirmar." % importe.group(1))

# ---------- Procesar correo del BANCO BBVA (transferencia interbancaria) ----------
ES_MES = {'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,'julio':7,
          'agosto':8,'septiembre':9,'setiembre':9,'octubre':10,'noviembre':11,'diciembre':12}

def parse_bbva(body):
    """Extrae los datos de un aviso de transferencia de BBVA (clientes@bbva.mx).
    OJO: BBVA NO manda 'Clave de Rastreo' (solo 'Folio Internet') -> el cruce con ORAMI
    se hace luego por monto+fecha (ver casar_recarga)."""
    folio   = re.search(r"Folio Internet\s*:?\s*(\d+)", body)
    # Importe tolerante: BBVA a veces separa los centavos (superindice) -> permite espacios
    # alrededor del punto y el simbolo $ opcional. Ej: "Importe $ 32,000.00" / "32,000 .00".
    mimp    = re.search(r"Importe[^\d]*?([\d][\d,]*)(?:\s*\.\s*(\d{2}))?(?!\d)", body)
    if not (mimp and folio):
        return None
    entero = mimp.group(1).replace(",", "").replace(" ", "")
    monto  = float(entero + "." + (mimp.group(2) or "00"))
    importe_txt = "{:,.2f}".format(monto)
    fecha   = re.search(r"Fecha de operaci[oó]n\s*:?\s*(\d{1,2}\s+de\s+[A-Za-zÁÉÍÓÚáéíóúÑñ]+\s+de\s+\d{4},?\s*\d{1,2}:\d{2}:\d{2}\s*[AaPp]\.?\s*[Mm]\.?)", body)
    benef   = re.search(r"Nombre del beneficiario\s+(.+?)\s+Importe", body)
    titular = re.search(r"Titular de la cuenta de retiro\s+(.+?)\s+Banco", body)
    bancod  = re.search(r"Banco Destino\s+(.+?)\s+Cuenta", body)
    cuenta  = re.search(r"Cuenta de dep[oó]sito\s+([\*\dxX]+)", body)
    concepto= re.search(r"Concepto de pago\s+(.+?)\s+Folio", body)
    fh = fecha.group(1).strip() if fecha else ""
    dt = datetime.now()
    m = re.match(r"(\d{1,2})\s+de\s+([A-Za-zÁÉÍÓÚáéíóúÑñ]+)\s+de\s+(\d{4}),?\s*(\d{1,2}):(\d{2}):(\d{2})\s*([AaPp])", fh)
    if m:
        mo = ES_MES.get(m.group(2).lower(), datetime.now().month)
        hh = int(m.group(4)) % 12 + (12 if m.group(7).upper()=="P" else 0)
        try: dt = datetime(int(m.group(3)), mo, int(m.group(1)), hh, int(m.group(5)), int(m.group(6)))
        except Exception: dt = datetime.now()
    return {
        "folio": folio.group(1),
        "monto": monto,
        "importe_txt": importe_txt,
        "dt": dt, "fechaHora": fh,
        "beneficiario": benef.group(1).strip() if benef else "CONSULTORIA EMPRESARIAL ORAMI S.C.",
        "titular": titular.group(1).strip() if titular else "MIGUEL ANGEL VALLEJO FLORES",
        "bancoDestino": bancod.group(1).strip() if bancod else "BAJIO",
        "cuenta": cuenta.group(1).strip() if cuenta else "",
        "concepto": concepto.group(1).strip() if concepto else "ORAMI",
    }

def procesar_bbva(body):
    """Devuelve True SOLO si creo una recarga nueva (para reenviar/notificar una sola vez)."""
    d = parse_bbva(body)
    if not d:
        log("correo BBVA sin importe/folio, ignorado"); return False
    doc_id = "rec-" + d["folio"]
    if ya_archivado(doc_id): return False
    ref_doc = db.collection("movimientos").document(doc_id)
    dt = d["dt"]
    orden = (dt - datetime(1899,12,30)).total_seconds()/86400.0
    MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    payload = {
        "tipo":"abono","servicio":"Recarga (transferencia)",
        "fecha":"%d-%s"%(dt.day, MESES[dt.month-1]),
        "fechaFull":"%d-%s-%d"%(dt.day, MESES[dt.month-1], dt.year),
        "hora":dt.strftime("%H:%M:%S"),"recibo":d["folio"],
        "banco":"SPEI a ORAMI (BBVA) - Folio "+d["folio"],
        "transferencia":d["monto"],"estado":"pendiente","orden":orden,"folioBanco":d["folio"],
        "comprobante":{
            "banco":"BBVA México","operacion":"Transferencia interbancaria SPEI (mismo día)",
            "fechaHora":d["fechaHora"],"ordenante":d["titular"],
            "beneficiario":d["beneficiario"],"clabe":d["cuenta"],"bancoDestino":d["bancoDestino"],
            "importe":"$%s MN"%d["importe_txt"],"referencia":d["folio"],"concepto":d["concepto"],
            "aplicacion":d["fechaHora"].split(",")[0] if d["fechaHora"] else "","clave":"(BBVA no envía clave de rastreo)"
        }
    }
    snap = ref_doc.get()
    if snap.exists:
        cur = snap.to_dict()
        # Auto-correccion: si el monto/fecha guardados difieren de lo leido ahora (p.ej. se
        # creo con un parser viejo que leyo mal el importe), lo corrige sin duplicar.
        if round(float(cur.get("transferencia",0) or 0),2) != round(d["monto"],2):
            payload["estado"] = cur.get("estado","pendiente")   # no revertir si ORAMI ya la verifico
            ref_doc.set(payload, merge=True)
            log("recarga BBVA CORREGIDA:", d["folio"], d["monto"])
        else:
            log("transferencia BBVA ya registrada:", d["folio"])
        return False
    ref_doc.set(payload)
    log("recarga BBVA creada:", d["folio"], d["monto"])
    enviar_push("Transferencia registrada", "Transferiste $%s a ORAMI. Recarga pendiente de confirmar." % d["importe_txt"])
    return True

# ---------- Reenviar el aviso de BBVA a un tercero (ORAMI) ----------
def _cuerpo_para_reenvio(msg):
    if msg.is_multipart():
        html=""; plain=""
        for part in msg.walk():
            if "attachment" in str(part.get("Content-Disposition")): continue
            ct=part.get_content_type()
            if ct not in ("text/html","text/plain"): continue
            try: payload=part.get_payload(decode=True).decode(part.get_content_charset() or "utf-8","ignore")
            except Exception: continue
            if ct=="text/html" and not html: html=payload
            elif ct=="text/plain" and not plain: plain=payload
        return ("html",html) if html else ("plain", plain or "(sin contenido)")
    try: payload=msg.get_payload(decode=True).decode(msg.get_content_charset() or "utf-8","ignore")
    except Exception: payload=str(msg.get_payload())
    return ("html" if msg.get_content_type()=="text/html" else "plain", payload or "(sin contenido)")

def _conectar_smtp():
    """Intenta SSL directo (465) y si falla, STARTTLS (587). Devuelve la sesion logueada."""
    errores = []
    for modo, puerto in (("ssl", SMTP_PORT), ("starttls", 587)):
        try:
            if modo == "ssl":
                s = smtplib.SMTP_SSL(SMTP_HOST, puerto, context=ssl.create_default_context(), timeout=30)
            else:
                s = smtplib.SMTP(SMTP_HOST, puerto, timeout=30)
                s.ehlo(); s.starttls(context=ssl.create_default_context()); s.ehlo()
            s.login(IMAP_USER, IMAP_PASS)
            return s
        except Exception as e:
            errores.append("%s:%s -> %s" % (modo, puerto, e))
    raise RuntimeError(" | ".join(errores))

def reenviar_aviso(destino, d):
    """Manda a ORAMI un AVISO PROPIO con los datos de la transferencia (no el correo de
    BBVA tal cual: el filtro del servidor lo clasifica como SPAM/phishing por traer el
    contenido del banco desde otro remitente — error 550). Devuelve True o el error."""
    try:
        texto = (
            "Aviso de transferencia realizada\n\n"
            "Se realizó una transferencia SPEI a %s.\n\n"
            "   Importe:        $%s MXN\n"
            "   Fecha:          %s\n"
            "   Folio Internet: %s\n"
            "   Concepto:       %s\n"
            "   Ordenante:      %s\n"
            "   Banco origen:   BBVA México\n\n"
            "Mensaje automático del sistema Control ORAMI de Nacionalíssimo.\n"
            "Cualquier duda, responder a este correo."
        ) % (d["beneficiario"], d["importe_txt"], d["fechaHora"] or d["dt"].strftime("%d/%m/%Y %H:%M"),
             d["folio"], d["concepto"], d["titular"])
        fwd = MIMEText(texto, "plain", "utf-8")
        fwd["Subject"] = "Transferencia a ORAMI - $%s - folio %s" % (d["importe_txt"], d["folio"])
        fwd["From"] = IMAP_USER
        fwd["To"] = destino
        fwd["Reply-To"] = IMAP_USER
        s = _conectar_smtp()
        s.sendmail(IMAP_USER, [destino], fwd.as_string())
        s.quit()
        log("aviso de transferencia enviado a", destino)
        return True
    except Exception as e:
        log("aviso a ORAMI fallo:", e)
        return str(e)[:200]

# ---------- Procesar recibo de FACEBOOK / META ----------
def es_facebook(subj, body):
    s=(subj or "").lower(); b=(body or ""); bl=b.lower()
    # remitente real de los recibos: noreply@business-updates.facebook.com (Meta for Business)
    if "business-updates.facebook.com" in bl or "facebookmail.com" in bl: return True
    if "Meta Platforms Ireland" in b: return True
    if "Amount billed" in b and "Reference number" in b: return True
    if "meta ads and marketing messages receipt" in s: return True
    if ("receipt" in s or "recibo" in s) and ("facebook" in s or "meta" in s): return True
    return False

EN_MES = {'Jan':1,'Feb':2,'Mar':3,'Apr':4,'May':5,'Jun':6,'Jul':7,'Aug':8,'Sep':9,'Oct':10,'Nov':11,'Dec':12}

def procesar_facebook(body):
    importe = re.search(r"Amount billed\s*MX\$?\s*([\d,]+\.\d{2})", body)
    ref     = re.search(r"Reference number\s+(?:i\s+)?([A-Z0-9]{6,})", body)
    fecha   = re.search(r"Invoice Date\s*([A-Z][a-z]{2}\s+\d{1,2},\s*\d{4}(?:,\s*\d{1,2}:\d{2}\s*[AP]M)?)", body)
    if not (importe and ref):
        log("recibo Facebook sin importe/referencia, ignorado"); return
    monto = round(float(importe.group(1).replace(",", "")), 2)
    rid = ref.group(1)
    doc_id = "fb-" + rid
    if ya_archivado(doc_id): return
    ref_doc = db.collection("movimientos").document(doc_id)
    if ref_doc.get().exists:
        log("recibo Facebook ya registrado:", rid); return
    MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    if fecha:
        p = re.match(r"([A-Z][a-z]{2})\s+(\d{1,2}),\s*(\d{4})(?:,\s*(\d{1,2}):(\d{2})\s*([AP])M)?", fecha.group(1))
        mo = EN_MES.get(p.group(1), datetime.now().month)
        hh = (int(p.group(4))%12 + (12 if p.group(6)=='P' else 0)) if p.group(4) else 0
        mm = int(p.group(5)) if p.group(5) else 0
        dt = datetime(int(p.group(3)), mo, int(p.group(2)), hh, mm)
    else:
        dt = datetime.now()
    orden = (dt - datetime(1899,12,30)).total_seconds()/86400.0
    ref_doc.set({
        "tipo":"cargo","servicio":"Facebook","origen":"facebook",
        "monto":monto,"banco":"Facebook (Meta) - Recibo "+rid,
        "fecha":"%d-%s"%(dt.day,MESES[dt.month-1]),
        "fechaFull":"%d-%s-%d"%(dt.day,MESES[dt.month-1],dt.year),
        "hora":dt.strftime("%H:%M:%S"),"recibo":rid,
        "orden":orden,"estado":"pendiente"
    })
    log("cargo Facebook capturado:", rid, monto)
    enviar_push("Facebook te cobró",
                "Facebook cobró $%s. Pendiente de confirmar con ORAMI." % importe.group(1))

# ---------- Procesar correo de ORAMI (xlsx) ----------
def parse_xlsx(data):
    ns = {'a':'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
    z = zipfile.ZipFile(io.BytesIO(data))
    shared = []
    if 'xl/sharedStrings.xml' in z.namelist():
        r = ET.fromstring(z.read('xl/sharedStrings.xml'))
        for si in r.findall('a:si', ns):
            shared.append(''.join(t.text or '' for t in si.findall('.//a:t', ns)))
    sheet = sorted([n for n in z.namelist() if n.startswith('xl/worksheets/sheet') and n.endswith('.xml')])[0]
    root = ET.fromstring(z.read(sheet))
    def ci(ref):
        L = ''.join(c for c in ref if c.isalpha()); idx=0
        for c in L: idx=idx*26+(ord(c)-64)
        return idx-1
    rows=[]
    for row in root.findall('.//a:row', ns):
        cells={}; mx=-1
        for c in row.findall('a:c', ns):
            v=c.find('a:v', ns); val=''
            if v is not None and v.text is not None:
                val = shared[int(v.text)] if c.get('t')=='s' else v.text
            k=ci(c.get('r')); cells[k]=val; mx=max(mx,k)
        rows.append([cells.get(i,'') for i in range(mx+1)])
    return rows

def servicio(d):
    d=d.upper()
    if 'SPEI RECIBIDO' in d: return 'Recarga (transferencia)'
    if 'FACEBK' in d or 'FACEBOOK' in d or 'METAPAY' in d: return 'Facebook'
    if 'SHOPIFY' in d: return 'Shopify'
    if 'ANTHROPIC' in d or 'CLAUDE' in d: return 'Claude (Anthropic)'
    if 'MICROSOFT' in d: return 'Microsoft'
    if 'MERCADOLIBRE' in d or 'MERPAGO' in d: return 'Mercado Libre'
    if 'RAILWAY' in d: return 'Railway'
    if 'ADOBE' in d: return 'Adobe'
    if 'MAILCHIMP' in d: return 'Mailchimp'
    return 'Otro'
def merchant(d):
    m=re.search(r'5161020004119535 (.+?) Tarjeta', d)
    return m.group(1).strip() if m else d[:50]

def casar_facebook(desc, monto, base, recibo):
    """Casa un cargo de Facebook del estado de cuenta de ORAMI con el recibo de Meta
    USANDO LA REFERENCIA (FACEBK *XXXXX en la descripcion), que es identica al
    'Reference number' del recibo de Meta -> el id del doc es fb-<ref>. Asi el cruce es
    100% confiable aunque varios cargos sean del mismo monto ($15,000). El monto YA NO se
    usa para emparejar (eso causaba duplicados).
    - Si ya existe fb-<ref> (recibo de Meta ya capturado): lo marca verificado por ORAMI.
    - Si NO existe (ORAMI llego antes que el recibo): crea el cargo bajo el id fb-<ref>
      para que el recibo posterior NO lo duplique (procesar_facebook ve que ya existe y se sale).
    Devuelve True si era un cargo FACEBK con referencia (manejado aqui); False si no la trae
    (p.ej. 'FACEBOOK MEXICO' o 'METAPAY' sin referencia -> sigue como cargo normal)."""
    m = re.search(r"FACEBK\s*\*?\s*([A-Z0-9]{6,})", (desc or "").upper())
    if not m:
        return False
    rid = m.group(1)
    ref_doc = db.collection("movimientos").document("fb-"+rid)
    if ref_doc.get().exists:
        ref_doc.update({"estado":"verificada","reciboOrami":recibo,"confirmadoOrami":True})
        log("cargo Facebook confirmado por ORAMI (ref %s):"%rid, recibo)
    else:
        if ya_archivado("fb-"+rid): return True   # pertenece a un mes ya cerrado, no recrear
        nb = dict(base)
        nb.update({"tipo":"cargo","monto":monto,"banco":"FACEBK *"+rid,"servicio":"Facebook",
                   "origen":"orami","estado":"verificada","confirmadoOrami":True,
                   "reciboOrami":recibo,"refFb":rid})
        ref_doc.set(nb, merge=True)
        log("cargo Facebook creado desde ORAMI bajo id de referencia:", rid)
    return True

def casar_facebook_por_monto(monto, serial, recibo_orami):
    """Para cargos de Facebook que ORAMI reporta SIN referencia (descripcion 'FACEBOOK MEXICO'
    o 'METAPAY'): los casa con un recibo de Meta (fb-<ref>) PENDIENTE del MISMO monto exacto y
    fecha cercana. Solo toca recibos PENDIENTES: los ya confirmados por su propia referencia
    FACEBK no se tocan, asi NO se mezclan dos cargos distintos del mismo monto (p.ej. dos de
    $4,922.78, uno FACEBK y otro FACEBOOK MEXICO = dos cargos reales). Idempotente por reciboOrami."""
    try:
        fbs = list(db.collection("movimientos").where("origen","==","facebook").stream())
    except Exception as e:
        log("query recibos fb fallo:", e); return False
    # idempotencia: si un recibo fb ya fue sellado con este recibo de ORAMI, ya se caso
    for d in fbs:
        if str(d.to_dict().get("reciboOrami","")) == str(recibo_orami):
            return True
    cand=[]
    for d in fbs:
        m=d.to_dict()
        if m.get("estado")=="pendiente" and abs(round(float(m.get("monto",0) or 0),2)-monto) < 0.01 and abs(float(m.get("orden",0))-serial) <= 7:
            cand.append((abs(float(m.get("orden",0))-serial), d))
    if not cand: return False
    cand.sort(key=lambda x:x[0])
    cand[0][1].reference.update({"estado":"verificada","confirmadoOrami":True,"reciboOrami":str(recibo_orami)})
    log("cargo Facebook (sin ref) confirmado por ORAMI (monto+fecha):", monto, "recibo", recibo_orami)
    return True

def casar_recarga(monto, serial, recibo_orami):
    """Casa un 'SPEI Recibido' del reporte de ORAMI con una recarga del banco.
    BBVA no manda clave de rastreo -> se empareja por monto bruto + fecha cercana.
    IDEMPOTENTE: sella la recarga casada con el recibo de ORAMI; si el reporte se
    REPROCESA (ORAMI lo reenvia o el motor corre otra vez), reconoce que esa linea ya
    se caso y NO crea un duplicado. Devuelve True si la linea ya quedo manejada."""
    try:
        abonos = list(db.collection("movimientos").where("tipo","==","abono").stream())
    except Exception as e:
        log("query recargas fallo:", e); return False
    # 1) idempotencia: ya hay una recarga sellada con este mismo recibo de ORAMI -> no dupliques
    for d in abonos:
        if str(d.to_dict().get("reciboOrami","")) == str(recibo_orami):
            return True
    # 2) casar por monto + fecha una recarga PENDIENTE, o una recarga del banco (BBVA, con
    #    folioBanco) que haya quedado verificada SIN sellar -> asi el reproceso no la duplica.
    cand=[]
    for d in abonos:
        m=d.to_dict()
        sellada = str(m.get("reciboOrami","")).strip()
        disponible = (m.get("estado")=="pendiente") or (m.get("folioBanco") and not sellada)
        if disponible and abs(round(float(m.get("transferencia",0) or 0),2)-monto) < 0.01 and abs(float(m.get("orden",0))-serial) <= 6:
            cand.append((abs(float(m.get("orden",0))-serial), d))
    if not cand: return False
    cand.sort(key=lambda x:x[0])
    cand[0][1].reference.update({"estado":"verificada","reciboOrami":str(recibo_orami)})
    log("recarga confirmada por ORAMI (monto+fecha):", monto, "recibo", recibo_orami)
    return True

def dedup_recargas():
    """Red de seguridad: si una recarga del banco (rec-, con folioBanco/comprobante) y un
    'SPEI recibido' de ORAMI (mov-) quedaron como DOS docs del mismo monto y dia (p.ej.
    porque el monto estuvo mal leido cuando llego el reporte), los fusiona: sella el rec-
    con el recibo de ORAMI, lo marca verificada y borra el mov-. Se emparejan 1 a 1
    (zip), asi dos transferencias reales del mismo monto/dia no se pierden."""
    try:
        abonos = list(db.collection("movimientos").where("tipo","==","abono").stream())
    except Exception as e:
        log("dedup: query fallo:", e); return
    grupos = {}
    for d in abonos:
        m = d.to_dict()
        key = (round(float(m.get("transferencia",0) or 0),2), str(m.get("fechaFull")))
        grupos.setdefault(key, []).append(d)
    for key, g in grupos.items():
        if len(g) < 2: continue
        bancos = [d for d in g if d.to_dict().get("folioBanco") or d.to_dict().get("comprobante")]
        movs   = [d for d in g if d.id.startswith("mov-") and not d.to_dict().get("folioBanco")]
        for b, mv in zip(bancos, movs):
            rec_orami = str(mv.to_dict().get("recibo",""))
            b.reference.update({"estado":"verificada","reciboOrami":rec_orami})
            mv.reference.delete()
            log("dedup: recarga duplicada fusionada — conservo %s, borro %s (reciboOrami=%s)" % (b.id, mv.id, rec_orami))

def procesar_orami(data):
    rows = parse_xlsx(data)
    MESES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    nuevos=0; verif=0; verif_fb=0
    for r in rows:
        if not r or not str(r[0]).strip().isdigit(): continue
        serial=float(r[1]); dt=datetime(1899,12,30)+timedelta(days=serial)
        desc=r[4] if len(r)>4 else ''; cargo=r[5] if len(r)>5 else ''; abono=r[6] if len(r)>6 else ''
        recibo=str(r[3]) if len(r)>3 else ''
        base={"orden":serial,"servicio":servicio(desc),
              "fecha":"%d-%s"%(dt.day,MESES[dt.month-1]),
              "fechaFull":"%d-%s-%d"%(dt.day,MESES[dt.month-1],dt.year),
              "hora":(datetime(1899,12,30)+timedelta(days=float(r[2]))).strftime("%H:%M:%S") if len(r)>2 and r[2] else dt.strftime("%H:%M:%S"),
              "recibo":recibo}
        if cargo not in ('',None):
            monto=round(float(cargo),2)
            if servicio(desc)=='Facebook':
                # 1) casar por referencia FACEBK *xxx (lo mas confiable)
                if casar_facebook(desc, monto, base, recibo):
                    verif_fb+=1; continue
                # 2) sin referencia (FACEBOOK MEXICO / METAPAY): casar un recibo fb PENDIENTE por monto
                if casar_facebook_por_monto(monto, serial, recibo):
                    verif_fb+=1; continue
            doc_id="mov-"+recibo
            if ya_archivado(doc_id): continue
            base.update({"tipo":"cargo","monto":monto,"banco":merchant(desc),"estado":"verificada","origen":"orami"})
            db.collection("movimientos").document(doc_id).set(base, merge=True); nuevos+=1
        elif abono not in ('',None):
            montoab=round(float(abono),2); hecho=False
            # 1) Banorte: casar por Clave de Rastreo (si viene en la descripcion)
            cve = re.search(r'Clave de Rastreo:\s*([A-Z0-9]+)', desc)
            if cve:
                rec = db.collection("movimientos").document("rec-"+cve.group(1))
                if rec.get().exists:
                    rec.update({"estado":"verificada"}); verif+=1; hecho=True
            # 2) BBVA (sin clave): casar por monto bruto + fecha (idempotente por recibo de ORAMI)
            if not hecho and casar_recarga(montoab, serial, recibo):
                verif+=1; hecho=True
            # 3) Si no caso con ninguna recarga del banco, crear el abono desde ORAMI
            if not hecho:
                doc_id="mov-"+recibo
                if ya_archivado(doc_id): continue
                base.update({"tipo":"abono","transferencia":montoab,"banco":"SPEI recibido","estado":"verificada"})
                db.collection("movimientos").document(doc_id).set(base, merge=True); nuevos+=1
    log(f"ORAMI procesado: {nuevos} movimientos, {verif} recargas verificadas, {verif_fb} cargos Facebook confirmados")
    # Diagnostico: cuantas filas trae el reporte y las ultimas 6 (fecha + monto) para ver
    # hasta que dia llega ORAMI y confirmar que se lee bien.
    try:
        datarows = [r for r in rows if r and str(r[0]).strip().isdigit()]
        ultimas = []
        for r in datarows[-6:]:
            dtt = datetime(1899,12,30)+timedelta(days=float(r[1]))
            monto = (r[5] if len(r)>5 and r[5] not in ('',None) else (r[6] if len(r)>6 else ''))
            ultimas.append("%d-%s %s %s" % (dtt.day, MESES[dtt.month-1], str(monto), (str(r[4])[:18] if len(r)>4 else "")))
        db.collection("tokens").document("zdebug_orami").set({
            "ts": ahora_mx().strftime("%Y-%m-%d %H:%M:%S"),
            "filas": len(datarows), "nuevos": nuevos, "verif": verif, "verif_fb": verif_fb,
            "ultimas": ultimas})
    except Exception as e:
        log("debug orami fallo:", e)
    # Notificar SOLO si hubo algo nuevo (el reporte se reprocesa cada corrida por el escaneo
    # de 7 dias; sin este candado mandaria un push repetido en cada corrida).
    if nuevos or verif or verif_fb:
        extra = f" y {verif_fb} cargos de Facebook confirmados" if verif_fb else ""
        enviar_push("Estado de cuenta de ORAMI",
                    f"Llegó el reporte de ORAMI: {nuevos} movimientos y {verif} recargas verificadas{extra}.")

# ---------- Main ----------
def main():
    # 0) Cierre mensual (solo actua si ya cambio el mes). Si falla (p.ej. el correo del
    #    respaldo), NO se borra nada y se reintenta en la siguiente corrida.
    try:
        cfg_c = cierre_mensual()
        try:
            db.collection("tokens").document("zdebug_cierre").set({
                "ts": ahora_mx().strftime("%Y-%m-%d %H:%M:%S"), "ok": True,
                "mesActual": cfg_c.get("mesActual",""), "saldoInicial": cfg_c.get("saldoInicial",0)})
        except Exception: pass
    except Exception as e:
        log("cierre mensual fallo (se reintenta en la proxima corrida):", e)
        try:
            db.collection("tokens").document("zdebug_cierre").set({
                "ts": ahora_mx().strftime("%Y-%m-%d %H:%M:%S"), "ok": False, "error": str(e)[:400]})
        except Exception: pass
    cargar_archivados()   # ids de meses cerrados: jamas se re-crean
    # Conexion IMAP con REINTENTOS: si el servidor de correo tiene un bache momentaneo
    # (no responde), se reintenta en vez de tumbar toda la corrida. Solo falla de verdad
    # si el correo esta caido en los 3 intentos.
    ctx = ssl.create_default_context()
    M = None
    for intento in range(3):
        try:
            M = imaplib.IMAP4_SSL(IMAP_HOST, IMAP_PORT, ssl_context=ctx, timeout=45)
            M.login(IMAP_USER, IMAP_PASS)
            M.select("INBOX")
            break
        except Exception as e:
            log("conexion IMAP fallo (intento %d/3): %s" % (intento+1, e))
            try:
                if M: M.logout()
            except Exception: pass
            M = None
            if intento < 2:
                import time as _t; _t.sleep(15)
    if M is None:
        log("no se pudo conectar al correo tras 3 intentos; se reintenta en la proxima corrida")
        return
    try:
        typ, data = M.uid('search', None, "UNSEEN")   # UID: estable aunque Outlook mueva correos
        ids = list(data[0].split())
    except Exception as e:
        log("busqueda de no-leidos fallo (se reintenta la proxima corrida):", e)
        try: M.logout()
        except Exception: pass
        return
    dbg = {"ts": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "unseen": len(ids), "banco": {}, "bbva": []}
    # Ademas: correos del BANCO y de ORAMI de los ultimos 7 dias AUNQUE ya esten leidos.
    # Outlook los marca leidos con la vista previa antes de que el motor (cada 15 min) los
    # alcance. Reprocesar es idempotente (rec-<folio>/rec-<clave>/mov-<recibo>/fb-<ref>), no
    # duplica. 'jgortizm' = quien manda el estado de cuenta de ORAMI (reportes xlsx).
    seen_set = set(ids)
    since = (datetime.now() - timedelta(days=7)).strftime("%d-%b-%Y")
    for addr in ("bbva.mx", "banorte", "jgortizm"):
        try:
            typ, d = M.uid('search', None, 'FROM', addr, 'SINCE', since)
            hits = d[0].split()
            dbg["banco"][addr] = len(hits)
            for n in hits:
                if n not in seen_set:
                    ids.append(n); seen_set.add(n)
        except Exception as e:
            dbg["banco"][addr] = "err:"+str(e)[:80]
            log("busqueda de correos del banco fallo:", e)
    log(f"correos a revisar: {len(ids)} (no leidos + banco de los ultimos 3 dias)")
    dbg["nids"] = len(ids); dbg["vistos"] = []
    for num in ids:
        try:
            typ, d = M.uid('fetch', num, "(RFC822)")
            if not d or not d[0]:
                log("fetch vacio, se omite:", num); continue
            msg = email.message_from_bytes(d[0][1])
            frm = decode(msg.get("From","")).lower()
            subj = decode(msg.get("Subject",""))
            log("correo de:", frm, "| asunto:", subj)
            xlsx = get_xlsx_attachment(msg)
            body = get_body(msg)
            if len(dbg["vistos"])<20: dbg["vistos"].append("%s | xlsx=%s" % (frm[:32], xlsx is not None))
            es_bbva = ("bbva" in frm) or ("interbancaria" in subj.lower())
            es_banorte = ("banorte" in frm) or ("transferencia" in subj.lower() and "spei" in subj.lower())
            if "jgortizm" in frm or "gerardo ortiz" in frm:
                dbg.setdefault("orami_dbg", []).append({"subj": subj[:40], "tiene_xlsx": xlsx is not None, "ruta": "?"})
            if es_bbva:
                mfol = re.search(r"Folio Internet\s*:?\s*(\d+)", body)
                _ix = body.find("Importe")
                dbg["bbva"].append({"folio": mfol.group(1) if mfol else "NO_MATCH",
                                    "importeCtx": (body[_ix:_ix+130] if _ix>=0 else "NO_Importe")})
                procesar_bbva(body)
                # Reenviar a ORAMI con REINTENTO: se marca reenviadoOrami en la recarga solo
                # cuando el envio sale bien; mientras no, se reintenta en cada corrida (los
                # correos del banco se re-escanean 3 dias, asi que hay muchas oportunidades).
                if mfol:
                    rdoc = db.collection("movimientos").document("rec-" + mfol.group(1))
                    rsnap = rdoc.get()
                    if rsnap.exists and not rsnap.to_dict().get("reenviadoOrami"):
                        datos = parse_bbva(body)
                        res = reenviar_aviso(REENVIO_BBVA, datos) if datos else "sin datos parseables"
                        if res is True:
                            rdoc.update({"reenviadoOrami": True})
                            dbg["bbva"][-1]["reenvio"] = "ok"
                        else:
                            dbg["bbva"][-1]["reenvio"] = res
            elif es_banorte:
                procesar_banco(body)
            elif xlsx is not None:
                if dbg.get("orami_dbg"): dbg["orami_dbg"][-1]["ruta"] = "procesar_orami"
                procesar_orami(xlsx)
            elif "facebook.com" in frm or es_facebook(subj, body):
                procesar_facebook(body)
            else:
                if dbg.get("orami_dbg"): dbg["orami_dbg"][-1]["ruta"] = "IGNORADO (no reconocido)"
                log("correo no reconocido, se ignora")
        except Exception as e:
            if dbg.get("orami_dbg") and dbg["orami_dbg"] and dbg["orami_dbg"][-1].get("ruta")=="procesar_orami":
                dbg["orami_dbg"][-1]["error"] = str(e)[:180]
            log("ERROR con un correo (se omite, NO tumba la corrida):", e)
        try:
            M.uid('store', num, "+FLAGS", "\\Seen")
        except Exception as e:
            log("no se pudo marcar leido:", e)
    try:
        dedup_recargas()   # red de seguridad contra recargas dobles
    except Exception as e:
        log("dedup fallo:", e)
    try:
        db.collection("tokens").document("zdebug_motor").set(dbg)
    except Exception as e:
        log("no se pudo escribir debug:", e)
    try:
        M.logout()
    except Exception as e:
        log("logout fallo:", e)
    log("listo")

if __name__ == "__main__":
    main()
